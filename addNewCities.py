import tkinter as tk
import os
from openpyxl import load_workbook

if(os.path.isfile('Record.xlsx') == False):
    WB = Workbook()
    ws = WB.active
    WB.save("Record.xlsx")
wb = load_workbook('Record.xlsx')
sheet = wb.active

def submit(E1,v,v1):
        units=['f','c']
        vals=[E1.get(),units[v],v1]
        i=sheet.max_row+1
        for _,cols in enumerate(['A{}','D{}','E{}']):
            sheet[cols.format(i)]=vals[_]
        wb.save('Record.xlsx')
        print("Added city {}".format(E1.get()))
def ADDCity():
    addCity = tk.Tk()
    var = tk.IntVar(master=addCity)
    var1 = tk.IntVar(master=addCity)
    var.set("1")
    var1.set("1")
    addCity.title("Add City")
    label = tk.Label(addCity,text = "City Name",bg="red",font=("gothic", 20))
    E1 = tk.Entry(addCity,bd =5,font=("gothic", 20))
    label.grid(column=0,row=0)
    E1.grid(column=1,row=0)
    
    
    rLbel = tk.Label(addCity,text = "Temperature",font=("gothic", 20))
    rad1 = tk.Radiobutton(addCity,text='f',variable=var,value=0,font=("gothic", 20))
    rad2 = tk.Radiobutton(addCity,text='c',variable=var,value=1,font=("gothic", 20))     
    rLbel.grid(column=0,row=1)
    rad1.grid(column=1,row=1)
    rad2.grid(column=2,row=1)
    
    
    ULbel = tk.Label(addCity,text = "Update(0/1)?",font=("gothic", 20))
    Urad1 = tk.Radiobutton(addCity,text='0',variable=var1,value=0,font=("gothic", 20))
    Urad2 = tk.Radiobutton(addCity,text='1',variable=var1,value=1,font=("gothic", 20))
    ULbel.grid(column=0,row=2)
    Urad1.grid(column=1,row=2)
    Urad2.grid(column=2,row=2)
    
    
    sub_btn=tk.Button(addCity,text = 'Submit', command = lambda: submit(E1,var.get(),var1.get())) 
    sub_btn.grid(column=1,row=3)
    
    addCity.mainloop()
ADDCity()